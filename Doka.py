from lectura_xml import lector
from enviar_mail import enviarmail_sin_error
from enviar_mail import enviarmail_con_error
import openpyxl
from getpass import getuser
from datetime import datetime
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import pythoncom
import os
import win32com.client
from reporte import ingresarsap,meteteensap
import shutil

def doka_oyp():
    ######INTERFAZ GRAFICA#######
    user=getuser()
    root=Tk()
    dt = datetime.now()
    day = dt.day
    month = dt.month
    year = dt.year

    horaExc = datetime.now().strftime('%Y-%m-%d_%H')

    hoy = str(day) + "-" + str(month) + "-" + str(year)
    root.title("Facturación_DOKA_OyP")
    root.resizable(0,0)
    root.geometry('450x250+450+200'.format(400, 400))
    miFrame=Frame(root,width=1000)
    miFrame.pack()
    miFrame2=Frame(root,width=1000)
    miFrame2.pack()
    miFrame3=Frame(root,width=1000)
    miFrame3.pack()
    miFrame4=Frame(root,width=1000)
    miFrame4.pack()

    ######CAMPOS DE INTERFAZ#######
    id0=IntVar()
    cliente0=Entry(miFrame4,textvariable=id0,width=20)
    cliente0.grid(row=2,column=2,padx=10,pady=5)
    rutalabel1=Label(miFrame4,text="Fecha DOKA (AAMMDD)",fg="Black",font=('Bold 1',10))
    rutalabel1.grid(row=1,column=2,sticky="n",padx=0,pady=0)
    
    ######CAMPOS DE INTERFAZ#######
    usuariosap=StringVar()
    cliente1=Entry(miFrame,textvariable=usuariosap,width=20)
    cliente1.grid(row=3,column=1,padx=30,pady=5)
    rutalabel2=Label(miFrame,text="Usuario",fg="Black",font=('Bold 1',10))
    rutalabel2.grid(row=2,column=1,sticky="n",padx=0,pady=0)
    
    contrasegna=StringVar()
    cliente1=Entry(miFrame,textvariable=contrasegna,width=20,show="*")
    cliente1.grid(row=3,column=3,padx=30,pady=5)
    rutalabel2=Label(miFrame,text="Contraseña",fg="Black",font=('Bold 1',10))
    rutalabel2.grid(row=2,column=3,sticky="n",padx=0,pady=0)

    hora = StringVar()
    cliente1=Entry(miFrame2,textvariable=hora,width=20)
    cliente1.grid(row=4,column=5,padx=30,pady=5)
    rutalabel2=Label(miFrame2,text="Hora Corte SAP (HH:MM:SS)",fg="Black",font=('Bold 1',10))
    rutalabel2.grid(row=3,column=5,sticky="n",padx=0,pady=0)
    
    fecha_corte=StringVar()
    hora2 = hora.get()
    fecha2 = fecha_corte.get()
    cliente1=Entry(miFrame2,textvariable=fecha_corte,width=20)
    cliente1.grid(row=4,column=1,padx=30,pady=5)
    rutalabel2=Label(miFrame2,text="Fecha Corte SAP (DD.MM.AAAA)",fg="Black",font=('Bold 1',10))
    rutalabel2.grid(row=3,column=1,sticky="n",padx=0,pady=0)
    
    def leerexcel():
        ######LEER EL EXCEL#######
        # pathagenda="C:/Users/"+ user + "/Desktop/doka/doka_oyp.xlsm"
        hora2 = hora.get()
        fecha2 = fecha_corte.get()
        pathagenda="C:/Users/"+ user + "/Desktop/doka/doka_oyp_"+str(fecha2[0:2])+ "_" +str(hora2[0:2])+"_hs.xlsm"
        print(pathagenda)
        wb = openpyxl.load_workbook(pathagenda,data_only=True,keep_vba=True)
        ws = wb["DETALLE ENVÍO DE MAIL"]
        ultimafiladelws=len(ws['A'])
        entregas=[]
        mail=[]
        nombre=[]
        asunto=[]
        texto=[]
        farmacia=[]
        resultado=[]
        fila=[]   
        
        ######TOMAR LOS DATOS DEL EXCEL#######
        for dato in range(2,ultimafiladelws+1):
            if dato==None:
                continue
            else:
                entregas.append(ws.cell(row=dato,column=1).value)
                mail.append(ws.cell(row=dato,column=2).value)
                nombre.append(ws.cell(row=dato,column=3).value)
                farmacia.append(ws.cell(row=dato,column=4).value)
                asunto.append(ws.cell(row=dato,column=5).internal_value)
                texto.append(ws.cell(row=dato,column=6).value)
                fila.append(ws.cell(row=dato,column=7).value)
                resultado.append(ws.cell(row=dato,column=8).value)
        wb.close()
        return entregas,mail,nombre,farmacia,asunto,texto,fila,resultado
    
    def generarexcel():
        hora2 = hora.get()
        fecha2 = fecha_corte.get()
        base=("C:/Users/"+ user + "/Desktop/doka/doka_oyp_base.xlsm")
        # shutil.copy(base, "C:/Users/"+ user + "/Desktop/doka/doka_oyp.xlsm")
        shutil.copy(base, "C:/Users/"+ user + "/Desktop/doka/doka_oyp_"+str(fecha2[0:2])+ "_" +str(hora2[0:2])+"_hs.xlsm")

    def cargardatos():
        hora2 = hora.get()
        fecha2 = fecha_corte.get()
        generarexcel()
        
        ingresarsap(usuariosap.get(),contrasegna.get())
        
        job = meteteensap(str(fecha_corte.get()), str(hora2), 0)

        if job == False:
            return messagebox.showinfo(title="Error!", message="No existen entregas para facturar. Por favor borrar el archivo generado recientemente.")
            
        pythoncom.CoInitialize()
        # path="C:/Users/"+ user + "/Desktop/doka/doka_oyp.xlsm"
        path="C:/Users/"+ user + "/Desktop/doka/doka_oyp_"+str(fecha2[0:2])+ "_" +str(hora2[0:2])+"_hs.xlsm"
        if os.path.exists(path):
            pythoncom.CoInitialize()
            Excel_macro = win32com.client.DispatchEx("Excel.Application") # DispatchEx is required in the newest versions of Python.
            Excel_path = os.path.expanduser(path)
            workbook = Excel_macro.Workbooks.Open(Filename = Excel_path, ReadOnly =1)
            # Excel_macro.Application.Run("doka_oyp.xlsm" + "!" + "z_integrador_doka.facturacion") # update Module1 with your module, Macro1 with your macro
            Excel_macro.Application.Run("doka_oyp_"+str(fecha2[0:2])+ "_" +str(hora2[0:2])+"_hs.xlsm" + "!" + "z_integrador_doka.facturacion") # update Module1 with your module, Macro1 with your macro
            workbook.Save()
            workbook.Close()
            Excel_macro.Application.Quit()
            del Excel_macro
            # messagebox.showinfo(message="Mensaje", title="Título")
            messagebox.showinfo(title = "Guardado Excel", message = "Proceso terminado. Gracias por aguardar.") 
    
    def mail():
        
        ######CAMPOS DE INTERFAZ#######
        # rutahtml = "C:/Users/" + user + "/Desktop/doka/ticket.html"
        dia=id0.get()
        
        ######UNPACKING DE COLUMNAS#######
        entregas,mail,nombre,farmacia,asunto,texto,fila,resultado=leerexcel()  
        cantidadenetregas=len(entregas)
        
        ######CARGAR EXCEL#######
        # pathagenda="C:/Users/"+ user + "/Desktop/doka/doka_oyp.xlsm"
        hora2 = hora.get()
        fecha2 = fecha_corte.get()
        pathagenda="C:/Users/"+ user + "/Desktop/doka/doka_oyp_"+str(fecha2[0:2])+ "_" +str(hora2[0:2])+"_hs.xlsm"
            
        wb = openpyxl.load_workbook(pathagenda,keep_vba=True)
        ws = wb["DETALLE ENVÍO DE MAIL"]
          
        ######PROCESO POR LINEA#######
        for i in range(0,cantidadenetregas):
            ruta="U:/Farmacias/DOKA_TICKETS_COBERTURA_PARCIAL"
            rutahtml=ruta+"/"+str(entregas[i])+".html"
            
            if mail[i]=="NOUSAR@SCIENZA.COM.AR" or mail[i]=="" or mail[i]==None or mail[i]=="NOUSAR@SCIENZA.COM" or mail[i]==" " or mail[i]=="NOUSA@SCIENZA.COM.AR" or mail[i]=="NOUSA@SCIENZA.COM": ######CONDICIONES DE ERROR DEL MAIL#######
                
                ######REEMPLAZO DE VALORES A EXCEPCIPON DE ERROR#######
                mail[i]="serviciopostentrega@scienza.com.ar"
                asunto[i]="TICKET NO ENVIADO POR ERROR EN LOS DATOS DEL AFILIADO"
                texto[i]="Por favor revisar mail del afiliado y reenviar ticket correspondiente a entrega: " + str(entregas[i])
                resultado[i]="ERROR ADV - Aviso de error enviado a Adm. de Ventas"

                try:
                    ######TRY EN CASO DE ERROR EN EL MAIL PARA ADM DE VENTAS#######
                    respuesta=lector(str(entregas[i]),rutahtml,dia,farmacia[i])
                    if respuesta == "descargado":
                        enviarmail_sin_error("Scienza1","serviciopostentrega@scienza.com.ar","serviciopostentrega@scienza.com.ar",asunto[i],texto[i],rutahtml,str(entregas[i])+".html")
                    
                    else:
                        ######TRY EN CASO DE ERROR EN LA IMPRESORA Y EN LOS DATOS DEL AFILIADO PARA ADM DE VENTAS#######
                        texto[i]="Por favor revisar mail del afiliado. Además, el ticket no fue encontrado. Entrega: " + str(entregas[i])
                        resultado[i]="ERROR FARMACIA / ADV - Ticket no encontrado. Aviso de error enviado a Adm. de Ventas"
                        enviarmail_con_error("Scienza1","serviciopostentrega@scienza.com.ar","serviciopostentrega@scienza.com.ar",asunto[i],texto[i],str(entregas[i])+".html")
                
                except:
                    ######TRY EN CASO DE ERROR EN LA IMPRESORA Y EN LOS DATOS DEL AFILIADO PARA ADM DE VENTAS#######
                    texto[i]="Por favor revisar mail del afiliado. Además, el ticket no fue encontrado. Entrega: " + str(entregas[i])
                    resultado[i]="ERROR FARMACIA / ADV - Ticket no encontrado. Aviso de error enviado a Adm. de Ventas"
                    enviarmail_con_error("Scienza1","serviciopostentrega@scienza.com.ar","serviciopostentrega@scienza.com.ar",asunto[i],texto[i],str(entregas[i])+".html")
            
            else:
                
                try:
                    ######TRY OK PARA ADM DE VENTAS#######
                    respuesta=lector(str(entregas[i]),rutahtml,dia,str(farmacia[i]))
                    
                    if respuesta == "descargado":
                        resultado[i]="OK - Ticket enviado"
                        enviarmail_sin_error("Scienza1","serviciopostentrega@scienza.com.ar",mail[i],asunto[i],texto[i],rutahtml,str(entregas[i])+".html")
                    
                    else:
                        ######TRY EN CASO DE ERROR EN LA IMPRESORA PARA ADM DE VENTAS#######
                        texto[i]="El ticket no fue encontrado. Entrega: " + str(entregas[i])
                        resultado[i]="ERROR FARMACIA - Ticket no encontrado"
                        enviarmail_con_error("Scienza1","serviciopostentrega@scienza.com.ar","serviciopostentrega@scienza.com.ar",asunto[i],texto[i],str(entregas[i])+".html")
                
                except:
                    ######TRY EN CASO DE ERROR EN LA IMPRESORA PARA ADM DE VENTAS#######
                    texto[i]="El ticket no fue encontrado. Entrega: " + str(entregas[i])
                    resultado[i]="ERROR FARMACIA - Ticket no generado"
                    enviarmail_con_error("Scienza1","serviciopostentrega@scienza.com.ar","serviciopostentrega@scienza.com.ar",asunto[i],texto[i],str(entregas[i])+".html")

            ws.cell(row=fila[i],column=8).value=resultado[i]
        # wb.save("C:/Users/"+ user + "/Desktop/doka/doka_oyp.xlsm")
        wb.save("C:/Users/"+ user + "/Desktop/doka/doka_oyp_"+str(fecha2[0:2])+ "_" +str(hora2[0:2])+"_hs.xlsm")
        wb.close()
        
        messagebox.showinfo(title = "Envío de mails", message = "Mails enviados. Por favor revisar resultados. Gracias por aguardar.")
    
    botonmail=Button(miFrame4, text="Mail",command=mail)
    botonmail.grid(row=2,column=3,sticky="e",padx=10,pady=10)

    botonarmado=Button(miFrame3, text="Inicio",command=cargardatos)
    botonarmado.grid(row=5,column=2,sticky="e",padx=10,pady=10)
    
    root.mainloop()
    
#---------------------------------------------------------fin---------------------------
# --------------------------------#
if __name__=="__main__":
    doka_oyp()