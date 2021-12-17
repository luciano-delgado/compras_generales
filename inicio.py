import openpyxl, tkinter as tk, numpy as np, pandas as pd
from getpass import getuser
from datetime import datetime
from tkinter import *
from tkinter import ttk
from tkinter import messagebox
from enviar_mail import enviarmail
from pruebas import generar_texto

def mail_facturas(mail_from,pass_from):
        
    user=getuser()
    pathagenda="C:/Users/"+ user + "/Desktop/mail_masivo/facturas_pendientes.xlsx"
    wb = openpyxl.load_workbook(pathagenda)
    ws = wb["Pendientes"]
    ws2 = wb["ListaDist"]
    ultimafiladelws=len(ws['A'])
    ultimafiladelws2=len(ws2['A'])
    l_fecha_derecepcion=[]
    l_proveedor=[]
    l_numero=[]
    l_fecha_de_emision=[]
    l_demora=[]
    l_fce=[]
    l_observaciones=[]
    l_fila=[]
    l_usuario=[]
    l_mail1_des=[]
    l_mail2_cc=[]
    l_mail3_cc=[]
    l_enviar_leido=[]
    #l_textos=[]   

    l_distribucion = []
    for valor in range(2,ultimafiladelws2+1):
        valor_ = str(ws2.cell(row=valor,column=1).value)
        if valor_ != 'None' and valor_ != None:
            l_distribucion.append(valor_)
    print("--------------------------------------------------------------------------------------------------------------")
    print(f' Lista Distribucion - facturasgenerales@scienza.com.ar: {l_distribucion}')
    print("--------------------------------------------------------------------------------------------------------------")

    for dato in range(2,ultimafiladelws+1): 
        if dato==None or dato =='None' or dato =='':
            continue
        else:
            l_fecha_derecepcion.append(ws.cell(row=dato,column=1).value)
            l_proveedor.append(ws.cell(row=dato,column=2).value) #uso
            l_numero.append(ws.cell(row=dato,column=3).value)   #uso
            l_fecha_de_emision.append(ws.cell(row=dato,column=4).value) #uso
            l_demora.append(ws.cell(row=dato,column=5).internal_value)
            l_fce.append(ws.cell(row=dato,column=6).value)
            l_observaciones.append(ws.cell(row=dato,column=7).value)
            l_usuario.append(ws.cell(row=dato,column=8).value)           #uso
            l_mail1_des.append(ws.cell(row=dato,column=9).value)
            l_mail2_cc.append(ws.cell(row=dato,column=10).value)
            l_fila.append(ws.cell(row=dato,column=11).value)
            l_enviar_leido.append(ws.cell(row=dato,column=12).value)
            #l_textos.append(ws.cell(row=dato,column=13).internal_value)
    
    usuario_anterior = 0
    lista_facturas = []
    lista_fe_emision = []
    lista_proveedores = []
    lista_usuario = []
    lista_enviar = []
    lista_enviar='Si'

    for i in range(0,len(l_fecha_derecepcion)): 
        usuario_leido = l_usuario[i]
        if usuario_leido == usuario_anterior or usuario_anterior == 0 or usuario_anterior=='':
            if l_enviar_leido[i] =="Si":
                if lista_enviar == 'Si':
                    lista_usuario.append(str(usuario_leido))
                    lista_facturas.append(str(l_numero[i]))
                    lista_fe_emision.append(str(l_fecha_de_emision[i]))
                    lista_proveedores.append(str(l_proveedor[i]))
                    usuario_anterior = usuario_leido
                else:   
                    lista_enviar = l_enviar_leido[i]
                            
        elif usuario_leido != usuario_anterior: # enviar mail si el usuario cambia. 
            print(f' Usuario: {set(lista_usuario)} - facturas a enviar: {len(lista_facturas)} [{lista_facturas}]')
            if lista_usuario != []:
                cuerpo_mail = generar_texto(lista_usuario[0], lista_facturas,lista_proveedores,lista_fe_emision,)
                l_cc = []
                l_cc.append(l_mail2_cc[i-1])
                l_cc_y_distribucion =  l_cc + l_distribucion
                enviarmail(pass_from, mail_from, l_mail1_des[i-1],l_cc_y_distribucion,"Recepcion Facturas",cuerpo_mail)     
                print(f' --- enviado a destinatario {l_mail1_des[i-1]} - responsable: {l_mail2_cc[i-1]} ')
            if l_usuario[i-1] != 'None' and l_usuario[i-1] != None and lista_facturas != 'None' and lista_facturas!= None:
                ws.cell(row=l_fila[i-1],column=13).value= f'Facturas enviadas a {l_usuario[i-1]}: {len(lista_facturas)} - nro fact: {lista_facturas}'
            if l_enviar_leido[i] == 'Si':
                usuario_anterior = l_usuario[i]
                lista_facturas = [str(l_numero[i])]
                lista_fe_emision = [l_fecha_de_emision[i]]
                lista_proveedores = [l_proveedor[i]]
                lista_usuario = [l_usuario[i]]
                lista_enviar = l_enviar_leido[i]
            else:
                usuario_anterior = l_usuario[i]
                lista_facturas = []
                lista_fe_emision = []
                lista_proveedores = []
                lista_usuario = []
                #lista_enviar = l_enviar_leido[i]

    wb.save(pathagenda)
    wb.close()
#-----------------------------------------------------------------------------------------------------------------------

#mail_facturas('ldelgado@scienza.com.ar','Chewie2019!0707')


######################################################################################################
########################################## INTERFAZ GRAFICA ##########################################
######################################################################################################
root = tk.Tk()
root.geometry('300x125')
root.title('OyP - Envio masivo - 2.0')
root.config(bg='blue')
label_from=tk.Label(root, text="Ingrese su mail y contraseña:", bg='lightgreen',width = 300 , height = 2 ,font =('calibri', 13))
label_from.pack()
mail_usuario, pass_usuario = tk.StringVar(), tk.StringVar()

input_mail = tk.Entry(root,textvariable=mail_usuario, width=75,bd=3,selectbackground='violet')
input_mail.pack()
input_pass = tk.Entry(root,textvariable=pass_usuario, width=75,bd=3,selectbackground='violet', show="*")
input_pass.pack()

boton_enviar = tk.Button(root,text="Iniciar envío masivo",command=lambda: mail_facturas(input_mail.get(),input_pass.get()),bg='lightblue',font =('calibri', 12)) 
boton_enviar.pack()
root.mainloop()



# ---------- Notas 
    #Version 2.0 para compilar 17.12
